"""
模块名称: cv_matcher/task.py
功能描述: 多线程执行候选人 AI 简历比对的核心入口。
"""

import pandas as pd
import openpyxl
from concurrent.futures import ThreadPoolExecutor, as_completed
from typing import Dict, Any, Optional

from src.core.logger import LoggerFactory
from src.modules.cv_matcher.utils import JSONCache
from src.modules.cv_matcher.processor import evaluate_candidate
from paths import DATA_DIR

logger = LoggerFactory.get_logger("cv_matcher_task")

# --- 路径定义 ---
SOURCE_TARGET_FILE = DATA_DIR / "Output" / "Get_result_leipin.xlsx"
CACHE_FILE = DATA_DIR / "Output" / "cv_match_cache.json"

def process_single_row(idx: int, jd: str, resume: str, cache_mgr: JSONCache) -> dict:
    """包装单行供多线程使用"""
    # jd or resume could be float NaN
    if pd.isna(jd):
        jd = ""
    if pd.isna(resume):
        resume = ""
        
    res = evaluate_candidate(str(jd), str(resume), cache_mgr)
    return {"index": idx, "result": res}

def run(params: Optional[Dict[str, Any]] = None) -> Dict[str, Any]:
    """
    标准插件入口
    """
    params = params or {}
    task_id = params.get("task_id", "cv_matcher")
    logger.info(f"[{task_id}] 🚀 cv_matcher(AI 简历深度匹配) 模块启动")

    try:
        # 1. 前置条件检查
        if not SOURCE_TARGET_FILE.exists():
            msg = f"❌ 无法执行比对，目标文件不存在: {SOURCE_TARGET_FILE}"
            logger.error(msg)
            return {"status": "error", "data": None, "message": msg}

        # 检查是否能在修改前抢占写入权限
        try:
            with open(SOURCE_TARGET_FILE, "a"): pass
        except PermissionError:
            msg = f"❌ 无法写入目标文件，请先关闭被占用文件: {SOURCE_TARGET_FILE}"
            logger.error(msg)
            return {"status": "error", "data": None, "message": msg}

        # 2. 读取主任务表及初始化缓存
        logger.info(f"[{task_id}] 📂 正在读取候选人数据清单: {SOURCE_TARGET_FILE}")
        df = pd.read_excel(str(SOURCE_TARGET_FILE))
        
        cache_mgr = JSONCache(CACHE_FILE)
        
        # 检查前置条件列
        if "简历描述对比" not in df.columns or "工作经历(全量)" not in df.columns:
            msg = "❌ 源文件缺少 '简历描述对比' 或 '工作经历(全量)' 其中一列，无法匹配"
            logger.error(msg)
            return {"status": "error", "data": None, "message": msg}

        # 3. 投递任务到线程池
        futures_map = {}
        results_map = {} # {row_idx: result_dict}
        
        # 根据需求定义 5 个并发线程
        max_workers = 5
        logger.info(f"[{task_id}] ⚙️ 开启线程池进行 AI 匹配运算 (并发: {max_workers})...")
        
        with ThreadPoolExecutor(max_workers=max_workers) as executor:
            for idx, row in df.iterrows():
                jd = row.get("简历描述对比")
                resume = row.get("工作经历(全量)")
                salary = row.get("岗位预算薪资", "")

                # 联动逻辑：将岗位预算薪资作为 JD 描述的一部分注入给大模型
                if pd.notna(salary) and str(salary).strip():
                    jd = f"{jd}\n\n【岗位预算薪资参考】: {salary}"

                future = executor.submit(process_single_row, idx, jd, resume, cache_mgr)
                futures_map[future] = idx
            
            # 收集结果
            completed = 0
            total = len(futures_map)
            for future in as_completed(futures_map):
                res_obj = future.result()
                results_map[res_obj["index"]] = res_obj["result"]
                completed += 1
                if completed % 10 == 0 or completed == total:
                    logger.info(f"[{task_id}] ⏳ 匹配进度: {completed}/{total}")

        # 4. OpenPyXL 写入处理 (由于只需要更新，严格使用 openpyxl 定位)
        logger.info(f"[{task_id}] 💾 正在将评估结果以增量注入的方式保存回 {SOURCE_TARGET_FILE}")
        wb = openpyxl.load_workbook(str(SOURCE_TARGET_FILE))
        sheet = wb.active
        
        # 提取表头寻找列位置
        header_vals = [cell.value for cell in sheet[1]] if getattr(sheet, 'max_row', 0) >= 1 else []
        
        match_pct_idx = None
        match_str_idx = None
        
        if "匹配度" in header_vals:
            match_pct_idx = header_vals.index("匹配度") + 1
        else:
            match_pct_idx = len(header_vals) + 1
            sheet.cell(row=1, column=match_pct_idx, value="匹配度")
            header_vals.append("匹配度")

        if "匹配结构" in header_vals:
            match_str_idx = header_vals.index("匹配结构") + 1
        else:
            match_str_idx = len(header_vals) + 1
            sheet.cell(row=1, column=match_str_idx, value="匹配结构")
            header_vals.append("匹配结构")

        # 将字典数据合并
        for pd_idx, ai_dict in results_map.items():
            excel_row = pd_idx + 2  # Pandas 0-indexed +1, Excel header +1
            
            pct_val = ai_dict.get("matching_percent", "")
            
            struct_val = ai_dict.get("matching_structure", {})
            summary_val = ai_dict.get("summary", "")
            
            # Stringify formatting
            assembled_str_val = f"【结构评分】\n{struct_val}\n\n【综合评价】\n{summary_val}"
            
            sheet.cell(row=excel_row, column=match_pct_idx, value=str(pct_val))
            sheet.cell(row=excel_row, column=match_str_idx, value=assembled_str_val)

        wb.save(str(SOURCE_TARGET_FILE))
        logger.info(f"[{task_id}] ✅ cv_matcher 执行完毕！")

        return {
            "status": "success",
            "data": {
                "processed": len(results_map),
                "output_path": str(SOURCE_TARGET_FILE)
            },
            "message": f"简历比对完毕，共处理 {len(results_map)} 行候选人数据"
        }

    except Exception as e:
        logger.exception(f"[{task_id}] 运行期间捕获异常: {e}")
        return {"status": "error", "data": None, "message": str(e)}
