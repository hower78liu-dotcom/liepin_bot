"""
模块名称: result_data_enricher/task.py
功能描述: 读取爬虫结果文件并与任务清单 test_search 联表获取需求对比数据。利用 openpyxl 回灌不损伤文件格式。
"""

import pandas as pd
import openpyxl
from pathlib import Path
from typing import Dict, Any, Optional

from src.core.logger import LoggerFactory
from src.modules.result_data_enricher.processor import process
from paths import DATA_DIR, CONFIG_DIR, LOG_DIR

logger = LoggerFactory.get_logger("result_data_enricher")

# --- 路径定义 ---
SOURCE_FILE = DATA_DIR / "Input" / "Liepin" / "test_search.xlsx"
TARGET_FILE = DATA_DIR / "Output" / "Get_result_leipin.xlsx"


def run(params: Optional[Dict[str, Any]] = None) -> Dict[str, Any]:
    """
    标准插件入口 — 被 main.py Orchestrator 动态调用

    :param params: 可选参数字典
    :return: 标准结果字典
    """
    params = params or {}
    task_id = params.get("task_id", "enricher")
    logger.info(f"[{task_id}] 🚀 result_data_enricher 模块启动")

    try:
        # 1. 存在性检查
        for file_path, name in [(SOURCE_FILE, "来源"), (TARGET_FILE, "目标")]:
            if not file_path.exists():
                msg = f"❌ {name}文件不存在: {file_path}"
                logger.error(msg)
                return {"status": "error", "data": None, "message": msg}

        # 2. 目标文件写入锁定检查
        try:
            with open(TARGET_FILE, "a"):
                pass
        except PermissionError:
            msg = f"❌ 无法写入目标文件，请先关闭被占用文件: {TARGET_FILE}"
            logger.error(msg)
            return {"status": "error", "data": None, "message": msg}

        # 3. 读取数据帧
        logger.info(f"[{task_id}] 📂 正在使用 Pandas 读取并分析数据集...")
        df_source = pd.read_excel(str(SOURCE_FILE))
        df_target = pd.read_excel(str(TARGET_FILE))

        # 4. 执行 Left Join 把 "简历描述对比" 映射上去
        logger.info(f"[{task_id}] 🔄 执行 DataFrame 内存级复合主键联接 (职位名称+公司名称+工作城市)...")
        df_merged, matched_cnt, unmatched_cnt = process(df_target, df_source)
        logger.info(f"[{task_id}] 📊 数据合并诊断结论: 成功匹配 {matched_cnt} 行 | 未匹配(置空) {unmatched_cnt} 行")

        # 5. 回灌到 Excel (维持原有格式)
        logger.info(f"[{task_id}] 💾 正在通过 OpenPyXL 安全同步写回至: {TARGET_FILE}")
        wb = openpyxl.load_workbook(str(TARGET_FILE))
        sheet = wb.active
        
        # 定位标题行并查找/初始化目标列索引
        header_row = [str(cell.value).strip() if cell.value else "" for cell in sheet[1]] if getattr(sheet, 'max_row', 0) >= 1 else []
        
        # --- 目标字段映射与索引定位 ---
        # 简历描述对比
        desc_col_idx = header_row.index("简历描述对比") + 1 if "简历描述对比" in header_row else None
        if not desc_col_idx:
            desc_col_idx = len(header_row) + 1
            sheet.cell(row=1, column=desc_col_idx, value="简历描述对比")
            header_row.append("简历描述对比")

        # 岗位预算薪资 (为 CV Matcher 铺路：该字段将在后续模块中作为 jd_content 的一部分注入给大模型)
        salary_col_idx = header_row.index("岗位预算薪资") + 1 if "岗位预算薪资" in header_row else None
        if not salary_col_idx:
            salary_col_idx = len(header_row) + 1
            sheet.cell(row=1, column=salary_col_idx, value="岗位预算薪资")
            header_row.append("岗位预算薪资")
        else:
            # 防覆盖逻辑：记录现有列状态
            logger.info(f"[{task_id}] 📝 目标表中检测到已存在 '岗位预算薪资' 列，将执行覆盖更新。")

        # 6. 流式同步合并后的数据
        # 利用 df_merged 的数据进行安全写回
        for idx in range(len(df_merged)):
            row_excel = idx + 2
            
            # 简历描述对比
            val_desc = df_merged.at[idx, '简历描述对比']
            sheet.cell(row=row_excel, column=desc_col_idx, value=val_desc)
            
            # 岗位预算薪资 (无损对接格式)
            if '岗位预算薪资' in df_merged.columns:
                val_salary = df_merged.at[idx, '岗位预算薪资']
                sheet.cell(row=row_excel, column=salary_col_idx, value=val_salary)

        wb.save(str(TARGET_FILE))
        logger.info(f"[{task_id}] ✅ result_data_enricher 执行顺利完成！")

        return {
            "status": "success",
            "data": {
                "matched": matched_cnt,
                "unmatched": unmatched_cnt,
                "output_path": str(TARGET_FILE)
            },
            "message": f"合并写回完成: 成功 {matched_cnt} | 失败 {unmatched_cnt}"
        }

    except Exception as e:
        logger.exception(f"[{task_id}] 运行崩溃: {e}")
        return {"status": "error", "data": None, "message": str(e)}
