"""
模块名称: liepin_search_preparer
功能描述: 将 Data Cleaner 输出的清洗后数据，同步覆写到 Liepin bot 搜刮模块所需的 Input (test_search.xlsx) 模板中。
"""

import pandas as pd
import openpyxl
from pathlib import Path
from typing import Dict, Any, Optional

from src.core.logger import LoggerFactory
from config.config_loader import Config
from paths import DATA_DIR, CONFIG_DIR, LOG_DIR

logger = LoggerFactory.get_logger("liepin_search_preparer")

# --- 路径定义 ---
SOURCE_FILE = DATA_DIR / "Output" / "DataCleaner" / "Data_Cleaned.xlsx"
TARGET_DIR = DATA_DIR / "Input" / "Liepin"
TARGET_FILE = TARGET_DIR / "test_search.xlsx"

# --- 字段映射协议 ---
COLUMNS_MAPPING = {
    "公司名称": "公司名称",
    "搜索关键词/摘要": "职位名称",
    "工作城市": "工作城市",
    "经验要求": "经验要求",
    "学历要求": "学历要求",
    "年龄要求": "年龄要求",
    "性别要求": "性别要求",
    "薪资范围 (月)": "薪资范围",
    "简历描述对比": "简历描述对比",
    "活跃度": "活跃度"
}


def run(params: Optional[Dict[str, Any]] = None) -> Dict[str, Any]:
    """
    标准插件入口 — 被 main.py Orchestrator 动态调用

    :param params: 可选参数字典
    :return: 标准结果字典
    """
    params = params or {}
    task_id = params.get("task_id", "search_prep")
    logger.info(f"[{task_id}] 🚀 liepin_search_preparer 模块启动")

    try:
        # 1. 自动生成依赖目录
        TARGET_DIR.mkdir(parents=True, exist_ok=True)

        # 2. Check source file existence
        if not SOURCE_FILE.exists():
            msg = f"❌ 源文件不存在: {SOURCE_FILE}"
            logger.error(msg)
            return {"status": "error", "data": None, "message": msg}

        # 3. Target File Write Permission Check (防止被 Excel 锁定)
        if TARGET_FILE.exists():
            try:
                with open(TARGET_FILE, "a"):
                    pass
            except PermissionError:
                msg = f"❌ 无法写入目标文件，请先关闭被占用文件: {TARGET_FILE}"
                logger.error(msg)
                return {"status": "error", "data": None, "message": msg}

        # 4. 读取源数据
        logger.info(f"[{task_id}] 📂 读取源数据: {SOURCE_FILE}")
        df_source = pd.read_excel(str(SOURCE_FILE))

        # 4.1 核心修复：对齐列名并执行预检
        df_source.columns = [str(c).strip() for c in df_source.columns]

        if "搜索关键词/摘要" not in df_source.columns:
            logger.error(f"[{task_id}] 源文件中缺失关键列：'搜索关键词/摘要'，请检查 DataCleaner 输出。")
            return {"status": "error", "message": "Source column missing"}

        # 5. 加载/初始目标 Workbook
        if TARGET_FILE.exists():
            wb = openpyxl.load_workbook(str(TARGET_FILE))
            if "Search_list" not in wb.sheetnames:
                wb.create_sheet("Search_list")
        else:
            wb = openpyxl.Workbook()
            # Rename default sheet
            sheet = wb.active
            sheet.title = "Search_list"

        sheet = wb["Search_list"]

        # 6. 处理表头动态定位与清空旧数据
        # 检查是否已有标题行
        header_row = [cell.value for cell in sheet[1]] if getattr(sheet, 'max_row', 0) >= 1 else []
        
        # 如果第一行是空，或者连"职位名称"这样的基础字段都没有，则重新写入表头
        if "职位名称" not in header_row:
            target_headers = ["职位名称", "公司名称", "工作城市", "经验要求", "学历要求", "年龄要求", "性别要求", "薪资范围", "搜索关键词/摘要", "简历描述对比", "活跃度"]
            for col_idx, col_name in enumerate(target_headers, start=1):
                sheet.cell(row=1, column=col_idx, value=col_name)
            header_row = target_headers
        else:
            # 清空从第 2 行起的所有已有数据内容
            max_r = sheet.max_row
            max_c = sheet.max_column
            if max_r >= 2:
                for r in range(2, max_r + 1):
                    for c in range(1, max_c + 1):
                        sheet.cell(row=r, column=c).value = None

        # 建立标题 -> 列索引(1-based)的映射关系
        target_header_to_col_idx = {val: idx for idx, val in enumerate(header_row, start=1) if val is not None}

        # 7. 数据同步遍历
        write_count = 0
        skip_count = 0
        
        # Determine current target row
        current_row = 2

        for idx, row in df_source.iterrows():
            # 核心修复：获取搜索词并处理空值填充与备选逻辑
            title_val = row.get("搜索关键词/摘要")
            
            # 备选逻辑：如果为空，尝试取 "职位名称 Target" (用户指示)
            if pd.isna(title_val) or str(title_val).strip() == "":
                title_val = row.get("职位名称 Target")
            
            # 若仍为空，记录 warning 并跳过该行
            if pd.isna(title_val) or str(title_val).strip() == "":
                logger.warning(f"[{task_id}] ⚠️ 源数据第 {idx + 2} 行'搜索关键词/摘要'及其备选字段均为空，跳过。")
                skip_count += 1
                continue

            # 写入 Mapping 中的列
            for src_col, tgt_col in COLUMNS_MAPPING.items():
                target_col_idx = target_header_to_col_idx.get(tgt_col)
                if not target_col_idx:
                    continue
                
                val = row.get(src_col)
                if pd.isna(val) or str(val).lower() == 'nan':
                    val = ""

                sheet.cell(row=current_row, column=target_col_idx, value=val)

            # 补齐 Search_list 专属字段：将搜索词同步填入“职位名称”和“搜索关键词/摘要”列
            for fallback_tgt in ["职位名称", "搜索关键词/摘要"]:
                idx_tgt = target_header_to_col_idx.get(fallback_tgt)
                if idx_tgt:
                    sheet.cell(row=current_row, column=idx_tgt, value=title_val)

            current_row += 1
            write_count += 1
        
        # 8. 保存执行结果
        wb.save(str(TARGET_FILE))
        logger.info(f"[{task_id}] ✅ 数据同步完成！成功写入 {write_count} 行，跳过 {skip_count} 行。结果文件: {TARGET_FILE}")

        return {
            "status": "success",
            "data": {
                "written": write_count,
                "skipped": skip_count,
                "output_path": str(TARGET_FILE)
            },
            "message": f"同步完成，写入 {write_count} 行数据"
        }

    except Exception as e:
        logger.exception(f"[{task_id}] 执行崩溃: {e}")
        return {"status": "error", "data": None, "message": str(e)}
